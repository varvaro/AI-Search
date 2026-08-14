"""PR7.5.2 — SAT Human Verification Protocol for Garáže ND Smíchov.

Turns benchmark/dataset/acceptance_nds_smichov.jsonl into a blank sign-off
sheet for a human reviewer, and nothing else. This module does not run
retrieval, does not call ai_search, and does not decide whether any case is
correct — it only re-projects facts that already exist in the dataset into a
form a person can review, print, or fill in on the site.

"Nevytvářej žádná nová fakta" is the operative constraint: every column is
either copied verbatim from the `BenchmarkCase` the dataset already declares,
or is a REQUIRED-BUT-EMPTY slot for the reviewer to write into. Nothing is
inferred by re-running the pipeline, and nothing is guessed when a field is
missing.

review_status derivation (the one non-trivial mapping in this file):

  human_verified=False                                      -> OPEN
  human_verified=True,  ground_truth_status == "verified"    -> VERIFIED
  human_verified=True,  ground_truth_status == "needs_review" -> FOUND_NOT_VERIFIED
  human_verified=True,  ground_truth_status == "unverified"   -> FAILED

This is a re-statement of two fields the dataset already carries
(`human_verified`, `ground_truth_status`), not a new judgment: a case can be
`human_verified=True` while its `ground_truth_status` still says
"needs_review" — that is exactly "a reviewer looked, found something, but has
not signed off the fact yet", which is what FOUND_NOT_VERIFIED means. Every
case in acceptance_nds_smichov.jsonl is `human_verified=False` today (see
PR7.5.1), so every row this module produces right now is OPEN — the protocol
is meant to be filled in, not to report a result that does not exist yet.
"""
from __future__ import annotations

import csv
from dataclasses import asdict, dataclass, field
from pathlib import Path

from .dataset.schema import (
    CRITICAL_CRITICALITY_NAMES,
    DATASET_DIR,
    BenchmarkCase,
    load_dataset,
)

NDS_DATASET = DATASET_DIR / "acceptance_nds_smichov.jsonl"

REVIEW_STATUS_OPEN = "OPEN"
REVIEW_STATUS_VERIFIED = "VERIFIED"
REVIEW_STATUS_FAILED = "FAILED"
REVIEW_STATUS_FOUND_NOT_VERIFIED = "FOUND_NOT_VERIFIED"
REVIEW_STATUSES = frozenset({
    REVIEW_STATUS_OPEN,
    REVIEW_STATUS_VERIFIED,
    REVIEW_STATUS_FAILED,
    REVIEW_STATUS_FOUND_NOT_VERIFIED,
})

COLUMNS = (
    "case_id",
    "category",
    "query",
    "criticality",
    "expected_documents",
    "verification_method",
    "expert_required",
    "human_verified",
    "reviewer_name",
    "review_date",
    "review_status",
    "review_comment",
)


@dataclass
class SatRow:
    case_id: str
    category: str
    query: str
    criticality: str
    expected_documents: str
    verification_method: str
    expert_required: bool
    human_verified: bool
    reviewer_name: str
    review_date: str
    review_status: str
    review_comment: str = ""

    def to_dict(self) -> dict:
        return asdict(self)


def derive_review_status(case: BenchmarkCase) -> str:
    """The only derived field in this module — see the module docstring."""
    if not case.human_verified:
        return REVIEW_STATUS_OPEN
    if case.ground_truth_status == "verified":
        return REVIEW_STATUS_VERIFIED
    if case.ground_truth_status == "unverified":
        return REVIEW_STATUS_FAILED
    return REVIEW_STATUS_FOUND_NOT_VERIFIED


def is_expert_required(case: BenchmarkCase) -> bool:
    """legal / financial / safety — the same set acceptance_metrics treats as
    a critical error class. Taken from the dataset's `criticality` field, not
    invented here."""
    return case.criticality in CRITICAL_CRITICALITY_NAMES


def case_to_row(case: BenchmarkCase) -> SatRow:
    return SatRow(
        case_id=case.id,
        category=case.category,
        query=case.question,
        criticality=case.criticality,
        expected_documents="; ".join(case.expected_documents),
        verification_method=case.verification_method,
        expert_required=is_expert_required(case),
        human_verified=case.human_verified,
        reviewer_name=case.verified_by,
        review_date=case.verification_date,
        review_status=derive_review_status(case),
        # Never seeded from case.notes: "review_comment" is the reviewer's own
        # remark, not the author's design note - conflating the two would make
        # it impossible to tell what the reviewer actually wrote.
        review_comment="",
    )


def load_cases(dataset_path: Path | str = NDS_DATASET) -> list[BenchmarkCase]:
    """Thin wrapper over the shared loader — kept here so a caller doesn't
    need to reach into benchmark.dataset.schema just to build a protocol."""
    return load_dataset(Path(dataset_path))


def build_sat_table(cases: list[BenchmarkCase]) -> list[SatRow]:
    """One row per case, in exactly the order the dataset file declares them.

    Order is never re-sorted: a reviewer working through a printed sheet
    top-to-bottom must land on the same case a colleague using the .jsonl
    directly would call case N.
    """
    return [case_to_row(case) for case in cases]


def export_csv(rows: list[SatRow], path: Path | str) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.to_dict())
    return path


def export_xlsx(rows: list[SatRow], path: Path | str) -> Path:
    """Same 12 columns as export_csv, plus light formatting so a reviewer
    opening it on the site can tell an OPEN critical row apart from a filled
    one at a glance. Formatting only - no cell carries a value the CSV export
    does not also carry.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SAT Protocol"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    expert_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    status_fills = {
        REVIEW_STATUS_OPEN: PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        REVIEW_STATUS_VERIFIED: PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        REVIEW_STATUS_FAILED: PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        REVIEW_STATUS_FOUND_NOT_VERIFIED: PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    }

    for col_index, name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=name)
        cell.font = header_font
        cell.fill = header_fill

    for row_index, row in enumerate(rows, start=2):
        data = row.to_dict()
        for col_index, name in enumerate(COLUMNS, start=1):
            value = data[name]
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            if name == "expert_required" and value:
                cell.fill = expert_fill
            if name == "review_status":
                cell.fill = status_fills.get(value, status_fills[REVIEW_STATUS_OPEN])
        sheet.cell(row=row_index, column=COLUMNS.index("query") + 1).alignment = Alignment(
            wrap_text=True,
        )

    widths = {
        "case_id": 22, "category": 18, "query": 48, "criticality": 12,
        "expected_documents": 42, "verification_method": 18, "expert_required": 14,
        "human_verified": 14, "reviewer_name": 18, "review_date": 14,
        "review_status": 20, "review_comment": 36,
    }
    for name, width in widths.items():
        sheet.column_dimensions[get_column_letter(COLUMNS.index(name) + 1)].width = width
    sheet.freeze_panes = "A2"

    workbook.save(path)
    return path


@dataclass
class SatProtocol:
    dataset_path: Path
    rows: list[SatRow] = field(default_factory=list)

    @property
    def case_count(self) -> int:
        return len(self.rows)

    @property
    def open_count(self) -> int:
        return sum(1 for r in self.rows if r.review_status == REVIEW_STATUS_OPEN)

    @property
    def expert_required_count(self) -> int:
        return sum(1 for r in self.rows if r.expert_required)


def generate_sat_protocol(
    dataset_path: Path | str = NDS_DATASET,
    *,
    csv_path: Path | str | None = None,
    xlsx_path: Path | str | None = None,
) -> SatProtocol:
    """Load the dataset, build the table, and write whichever exports were
    requested. Neither export is required - callers that only want the
    in-memory table can omit both paths."""
    dataset_path = Path(dataset_path)
    cases = load_cases(dataset_path)
    rows = build_sat_table(cases)
    if csv_path is not None:
        export_csv(rows, csv_path)
    if xlsx_path is not None:
        export_xlsx(rows, xlsx_path)
    return SatProtocol(dataset_path=dataset_path, rows=rows)


def main(argv: list[str] | None = None) -> int:
    import argparse

    parser = argparse.ArgumentParser(
        prog="python -m benchmark.sat_protocol",
        description="Generate the SAT human verification protocol (CSV + XLSX) "
                     "for an acceptance dataset.",
    )
    parser.add_argument("--dataset", default=str(NDS_DATASET))
    parser.add_argument("--out-dir", default=str(Path(__file__).parent / "reports" / "sat"))
    parser.add_argument("--csv-name", default="sat_protocol.csv")
    parser.add_argument("--xlsx-name", default="sat_protocol.xlsx")
    args = parser.parse_args(argv)

    out_dir = Path(args.out_dir)
    protocol = generate_sat_protocol(
        args.dataset,
        csv_path=out_dir / args.csv_name,
        xlsx_path=out_dir / args.xlsx_name,
    )
    print(f"SAT protocol for {Path(args.dataset).name}")
    print(f"  case count:            {protocol.case_count}")
    print(f"  expert_required rows:  {protocol.expert_required_count}")
    print(f"  OPEN rows:             {protocol.open_count}")
    print(f"  CSV:  {out_dir / args.csv_name}")
    print(f"  XLSX: {out_dir / args.xlsx_name}")
    return 0


if __name__ == "__main__":
    import sys

    sys.exit(main())
